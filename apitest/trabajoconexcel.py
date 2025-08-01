from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

# --- Creating a new Excel file ---
def create_excel_file(filename="pruebaExcel.xlsx"):
    wb = Workbook()  # Create a new workbook
    ws = wb.active   # Get the active worksheet
    ws.title = "Mi Hoja desde python" # Set the sheet title
    
    # Style header rows
    header_font = Font(bold=True, color="EFFFFF")
    header_fill = PatternFill("solid", fgColor="5981BD")
    header_align = Alignment(horizontal="center")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
   
   # Style data rows
    data_font = Font(color="740000")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.font = data_font
            cell.alignment = Alignment(horizontal="left")
    
    
    
    # Write data to cells
    ws['A1'] = "Nombre"
    ws['B1'] = "edad"
    ws['A2'] = "Marianos"
    ws['B2'] = 44
    ws['A3'] = "Mateos"
    ws['B3'] = 15

    wb.save(filename)
    print(f"'{filename}' created successfully.")


# --- Reading an existing Excel file ---
def read_excel_file(filename="mi_excel.xlsx"):
    try:
        wb = load_workbook(filename)  # Load an existing workbook
        ws = wb.active  # Get the active worksheet

        print(f"\nReading data from '{filename}':")
        for row in ws.iter_rows(min_row=1, values_only=True):
            print(row)
    except FileNotFoundError:
        print(f"Error: '{filename}' not found.")


# --- Modifying an existing Excel file ---
def modify_excel_file(filename="mi_excel.xlsx"):
    try:
        wb = load_workbook(filename)
        ws = wb.active

        # Modify a cell's value
        ws['B2'] = 31

        # Add a new row
        ws.append(["Charlie", 35])

        wb.save(filename)
        print(f"\n'{filename}' modified successfully.")
    except FileNotFoundError:
        print(f"Error: '{filename}' not found.")

# éste "IF" lo que hace es preguntar si la variable
#  "__name__" del scrip que se ejecuta con ej: "python modulo.py" se llama "__main__", y como eso
#  ocurre cada vez q se ejecuta el scrip con python le asina el name=main_

if __name__ == "__main__":
    create_excel_file()
    #read_excel_file()
    #modify_excel_file()
    #read_excel_file() # Read again to see modifications